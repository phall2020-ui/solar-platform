"""
AMPYR Distributed Energy - Report Generation
=============================================
Professional PDF and Excel report generation with AMPYR branding.
"""

import os
import sys
from datetime import date

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm, cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                 Spacer, PageBreak, HRFlowable, KeepTogether)
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from . import constants as C
from .engine import ModelInputs, ModelResults


# ── Brand Colours for ReportLab ──────────────────────────────────────

RL_NAVY = colors.HexColor(C.BRAND_NAVY)
RL_NAVY_LIGHT = colors.HexColor(C.BRAND_NAVY_LIGHT)
RL_NAVY_MID = colors.HexColor(C.BRAND_NAVY_MID)
RL_TEAL = colors.HexColor(C.BRAND_TEAL)
RL_WHITE = colors.HexColor(C.BRAND_WHITE)
RL_GREY = colors.HexColor(C.BRAND_GREY)
RL_GREY_LIGHT = colors.HexColor(C.BRAND_GREY_LIGHT)
RL_RED = colors.HexColor(C.BRAND_RED_WARNING)
RL_GREEN = colors.HexColor(C.BRAND_GREEN)

PAGE_W, PAGE_H = A4


# ── PDF Styles ───────────────────────────────────────────────────────

def _get_styles():
    styles = getSampleStyleSheet()

    styles.add(ParagraphStyle(
        "BrandTitle", parent=styles["Title"],
        fontSize=22, textColor=RL_WHITE, fontName="Helvetica-Bold",
        spaceAfter=2 * mm,
    ))
    styles.add(ParagraphStyle(
        "BrandSubtitle", parent=styles["Normal"],
        fontSize=11, textColor=RL_GREY_LIGHT, fontName="Helvetica",
        spaceAfter=4 * mm,
    ))
    styles.add(ParagraphStyle(
        "SectionHead", parent=styles["Heading2"],
        fontSize=13, textColor=RL_TEAL, fontName="Helvetica-Bold",
        spaceBefore=8 * mm, spaceAfter=3 * mm,
    ))
    styles.add(ParagraphStyle(
        "BodyText2", parent=styles["Normal"],
        fontSize=9, textColor=colors.black, fontName="Helvetica",
        leading=12,
    ))
    styles.add(ParagraphStyle(
        "Warning", parent=styles["Normal"],
        fontSize=9, textColor=RL_RED, fontName="Helvetica-Bold",
        spaceBefore=2 * mm,
    ))
    styles.add(ParagraphStyle(
        "Footer", parent=styles["Normal"],
        fontSize=7, textColor=RL_GREY, fontName="Helvetica",
        alignment=TA_CENTER,
    ))
    return styles


# ── Header / Footer Drawing ─────────────────────────────────────────

def _draw_header(canvas, doc):
    """Draw AMPYR branded header on each page."""
    canvas.saveState()

    # Navy header bar
    canvas.setFillColor(RL_NAVY)
    canvas.rect(0, PAGE_H - 22 * mm, PAGE_W, 22 * mm, fill=1, stroke=0)

    # Teal accent line
    canvas.setFillColor(RL_TEAL)
    canvas.rect(0, PAGE_H - 22.5 * mm, PAGE_W, 0.7 * mm, fill=1, stroke=0)

    # Teal vertical accent bar (logo chevron approximation)
    canvas.rect(15 * mm, PAGE_H - 19 * mm, 1.2 * mm, 14 * mm, fill=1, stroke=0)

    # AMPYR text
    canvas.setFillColor(RL_WHITE)
    canvas.setFont("Helvetica-Bold", 18)
    canvas.drawString(20 * mm, PAGE_H - 16 * mm, "AMPYR")

    # Divider
    canvas.setFillColor(RL_GREY)
    canvas.setFont("Helvetica", 18)
    canvas.drawString(52 * mm, PAGE_H - 16 * mm, "|")

    # DISTRIBUTED ENERGY
    canvas.setFillColor(RL_GREY_LIGHT)
    canvas.setFont("Helvetica", 9)
    canvas.drawString(58 * mm, PAGE_H - 13 * mm, "DISTRIBUTED")
    canvas.drawString(58 * mm, PAGE_H - 17 * mm, "ENERGY")

    # Right side: report title
    canvas.setFillColor(RL_TEAL)
    canvas.setFont("Helvetica", 9)
    canvas.drawRightString(PAGE_W - 15 * mm, PAGE_H - 15 * mm,
                           "Solar PPA Pricing Report")

    # Footer
    canvas.setFillColor(RL_GREY)
    canvas.setFont("Helvetica", 7)
    canvas.drawString(15 * mm, 10 * mm,
                      "AMPYR Distributed Energy  —  Confidential")
    canvas.drawRightString(PAGE_W - 15 * mm, 10 * mm,
                           f"Page {doc.page}")

    canvas.restoreState()


# ── Table Helper ─────────────────────────────────────────────────────

def _make_table(data, col_widths=None, header_row=True):
    """Create a branded table."""
    t = Table(data, colWidths=col_widths, repeatRows=1 if header_row else 0)

    style_cmds = [
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("GRID", (0, 0), (-1, -1), 0.3, RL_GREY),
    ]

    if header_row and len(data) > 0:
        style_cmds.extend([
            ("BACKGROUND", (0, 0), (-1, 0), RL_NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), RL_WHITE),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
        ])

    # Alternate row shading
    for i in range(1, len(data)):
        if i % 2 == 0:
            style_cmds.append(("BACKGROUND", (0, i), (-1, i),
                               colors.HexColor("#F5F7FA")))

    t.setStyle(TableStyle(style_cmds))
    return t


def _kv_table(pairs: list, col_widths=None):
    """Key-value pair table (no header row)."""
    if col_widths is None:
        col_widths = [70 * mm, 60 * mm]
    t = Table(pairs, colWidths=col_widths)
    style_cmds = [
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 0), (-1, -2), 0.3, colors.HexColor("#DEE2E6")),
    ]
    t.setStyle(TableStyle(style_cmds))
    return t


# ── PDF Generation ───────────────────────────────────────────────────

def generate_pdf(filepath: str, res: ModelResults, inp: ModelInputs):
    """Generate branded PDF pricing report."""
    styles = _get_styles()

    doc = SimpleDocTemplate(
        filepath, pagesize=A4,
        topMargin=28 * mm, bottomMargin=18 * mm,
        leftMargin=15 * mm, rightMargin=15 * mm,
    )

    story = []

    # ── Cover / Title ──
    story.append(Spacer(1, 8 * mm))
    story.append(Paragraph(inp.project_name, styles["BrandTitle"]))
    story.append(Paragraph(
        f"Solar PPA Pricing Report  —  Generated {date.today().strftime('%d %B %Y')}",
        styles["BrandSubtitle"]))
    story.append(HRFlowable(width="100%", thickness=0.5, color=RL_TEAL,
                             spaceAfter=6 * mm))

    # ── Warnings ──
    if res.warnings:
        story.append(Paragraph("Warnings", styles["SectionHead"]))
        for w in res.warnings:
            story.append(Paragraph(f"⚠ {w}", styles["Warning"]))

    # ── Key Results ──
    story.append(Paragraph("Key Results", styles["SectionHead"]))
    story.append(_kv_table([
        ["Solve Method", res.solve_method],
        ["Year 1 PPA Rate", f"{res.ppa_rate_year1_pence:.2f} p/kWh"],
        ["Development Fee", f"£{res.dev_fee_total_solved:,.0f}  (£{res.dev_fee_per_kwp_solved:,.1f}/kWp)"],
        ["Levered IRR", f"{res.levered_irr * 100:.2f}%"],
        ["Target IRR", f"{res.target_irr * 100:.2f}%"],
        ["Unlevered IRR", f"{res.unlevered_irr * 100:.2f}%"],
        ["Cash Multiple", f"{res.cash_multiple:.2f}x"],
        ["Payback Period", f"Year {res.payback_year}"],
        ["Average DSCR", f"{res.average_dscr:.2f}x"],
    ]))

    # ── Key Dates ──
    story.append(Paragraph("Key Dates", styles["SectionHead"]))
    story.append(_kv_table([
        ["Financial Close", res.financial_close.strftime("%d %b %Y")],
        ["Construction Start", res.construction_start.strftime("%d %b %Y")],
        ["COD", res.cod.strftime("%d %b %Y")],
        ["Operations End", res.operations_end.strftime("%d %b %Y")],
        ["First Indexation", res.first_indexation_date.strftime("%d %b %Y")],
        ["Senior Debt Maturity", res.debt_end.strftime("%d %b %Y")],
    ]))

    # ── Capex ──
    story.append(Paragraph("Capital Expenditure", styles["SectionHead"]))
    story.append(_kv_table([
        ["EPC Total", f"£{res.epc_total:,.0f}"],
        ["Insurance (Construction)", f"£{res.insurance_construction:,.0f}"],
        ["Transaction Costs", f"£{res.transaction_costs:,.0f}"],
        ["Development Fee", f"£{res.development_fee_total:,.0f}"],
        ["Management Company Fee", f"£{res.management_company_fee:,.0f}"],
        ["Stamp Duty", f"£{res.stamp_duty:,.0f}"],
        ["Contingency", f"£{res.contingency:,.0f}"],
        ["Total Capex", f"£{res.total_capex:,.0f}"],
    ]))

    # ── Funding ──
    story.append(Paragraph("Funding Structure", styles["SectionHead"]))
    story.append(_kv_table([
        ["Senior Debt", f"£{res.senior_debt_amount:,.0f}  ({inp.debt_gearing * 100:.0f}%)"],
        ["SHL", f"£{res.shl_amount:,.0f}"],
        ["Pure Equity", f"£{res.pure_equity:,.0f}"],
    ]))

    # ── Customer Savings ──
    story.append(Paragraph("Customer Savings Summary", styles["SectionHead"]))
    story.append(_kv_table([
        ["Year 1 Saving", f"£{res.year1_saving:,.0f}"],
        ["Average Annual Saving", f"£{res.avg_annual_saving:,.0f}"],
        ["Total Lifetime Saving", f"£{res.total_lifetime_saving:,.0f}"],
        ["Avg Lifetime PPA Rate", f"{res.avg_lifetime_ppa_rate:.2f} p/kWh"],
        ["Avg Lifetime Grid Rate", f"{res.avg_lifetime_grid_rate:.2f} p/kWh"],
    ]))

    # ── Page break before cashflows ──
    story.append(PageBreak())

    # ── Annual Cashflow Summary ──
    story.append(Paragraph("Annual Cashflow Summary", styles["SectionHead"]))

    cf_header = ["Year", "Generation\n(MWh)", "PPA Rate\n(p/kWh)", "Revenue\n(£)",
                 "Opex\n(£)", "EBITDA\n(£)", "CFADS\n(£)", "Debt Service\n(£)",
                 "FCF\n(£)", "DSCR"]
    cf_data = [cf_header]
    for cf in res.cashflows:
        cf_data.append([
            f"{cf.year}",
            f"{cf.generation_kwh / 1000:,.0f}",
            f"{cf.ppa_price_pence:.2f}",
            f"{cf.revenue:,.0f}",
            f"{cf.total_opex:,.0f}",
            f"{cf.ebitda:,.0f}",
            f"{cf.cfads:,.0f}",
            f"{cf.debt_service:,.0f}",
            f"{cf.fcf_to_equity:,.0f}",
            f"{cf.dscr:.2f}" if cf.dscr > 0 else "—",
        ])

    cf_widths = [12 * mm, 18 * mm, 16 * mm, 20 * mm, 18 * mm, 20 * mm,
                 20 * mm, 20 * mm, 18 * mm, 14 * mm]
    story.append(_make_table(cf_data, col_widths=cf_widths))

    # ── Page break before customer view ──
    story.append(PageBreak())

    # ── PPA vs Grid Rate (Customer View) ──
    story.append(Paragraph("PPA vs Grid Rate — Customer View", styles["SectionHead"]))

    cust_header = ["Year", "Generation\n(MWh)", "PPA Rate\n(p/kWh)", "PPA Cost\n(£)",
                   "Grid Rate\n(p/kWh)", "Grid Cost\n(£)", "Annual\nSaving (£)",
                   "Buyout\nValue (£)"]
    cust_data = [cust_header]
    for cf in res.cashflows:
        cust_data.append([
            f"{cf.year}",
            f"{cf.generation_kwh / 1000:,.0f}",
            f"{cf.ppa_price_pence:.2f}",
            f"{cf.revenue:,.0f}",
            f"{cf.grid_price_pence:.2f}",
            f"{cf.grid_cost:,.0f}",
            f"{cf.customer_saving:,.0f}",
            f"{cf.buyout_value:,.0f}",
        ])

    cust_widths = [12 * mm, 18 * mm, 18 * mm, 20 * mm, 18 * mm, 20 * mm,
                   22 * mm, 22 * mm]
    story.append(_make_table(cust_data, col_widths=cust_widths))

    # ── Page break for equity cashflows ──
    story.append(PageBreak())

    # ── Equity / SHL Cashflows ──
    story.append(Paragraph("Equity & SHL Cashflows", styles["SectionHead"]))

    eq_header = ["Year", "FCF to\nEquity (£)", "SHL\nInterest (£)",
                 "SHL\nRepay (£)", "Equity\nDist (£)", "Cum.\nDist (£)",
                 "Debt\nBalance (£)", "SHL\nBalance (£)"]
    eq_data = [eq_header]
    for cf in res.cashflows:
        eq_data.append([
            f"{cf.year}",
            f"{cf.fcf_to_equity:,.0f}",
            f"{cf.shl_interest:,.0f}",
            f"{cf.shl_repayment:,.0f}",
            f"{cf.equity_distribution:,.0f}",
            f"{cf.cumulative_distributions:,.0f}",
            f"{cf.debt_balance_end:,.0f}",
            f"{cf.shl_balance_end:,.0f}",
        ])

    eq_widths = [12 * mm, 22 * mm, 20 * mm, 20 * mm, 20 * mm, 22 * mm,
                 22 * mm, 22 * mm]
    story.append(_make_table(eq_data, col_widths=eq_widths))

    # ── Buyout Schedule ──
    story.append(Paragraph("Annual Buyout Schedule", styles["SectionHead"]))
    bo_data = [["Year", "Buyout Value (£)"]]
    for cf in res.cashflows:
        bo_data.append([f"Year {cf.year}", f"£{cf.buyout_value:,.0f}"])
    story.append(_make_table(bo_data, col_widths=[30 * mm, 50 * mm]))

    # ── Build PDF ──
    doc.build(story, onFirstPage=_draw_header, onLaterPages=_draw_header)


# ── Excel Generation ─────────────────────────────────────────────────

# Brand fills for Excel
_NAVY_FILL = PatternFill(start_color="0B1A2E", end_color="0B1A2E", fill_type="solid")
_TEAL_FILL = PatternFill(start_color="00B4A0", end_color="00B4A0", fill_type="solid")
_LIGHT_FILL = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
_WHITE_FONT = Font(color="FFFFFF", bold=True, size=10)
_NAVY_FONT = Font(color="0B1A2E", bold=True, size=10)
_TEAL_FONT = Font(color="00B4A0", bold=True, size=11)
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=9)
_BODY_FONT = Font(size=9)
_NUMBER_FORMAT_GBP = '£#,##0'
_NUMBER_FORMAT_GBP_2 = '£#,##0.00'
_NUMBER_FORMAT_PCT = '0.00%'
_NUMBER_FORMAT_PENCE = '0.00'
_THIN_BORDER = Border(
    left=Side(style="thin", color="DEE2E6"),
    right=Side(style="thin", color="DEE2E6"),
    top=Side(style="thin", color="DEE2E6"),
    bottom=Side(style="thin", color="DEE2E6"),
)


def _write_kv_section(ws, row, title, pairs):
    """Write a key-value section to worksheet."""
    ws.cell(row=row, column=1, value=title).font = _TEAL_FONT
    row += 1
    for key, val in pairs:
        ws.cell(row=row, column=1, value=key).font = Font(bold=True, size=9)
        cell = ws.cell(row=row, column=2, value=val)
        cell.font = _BODY_FONT
        cell.alignment = Alignment(horizontal="right")
        row += 1
    return row + 1


def generate_excel(filepath: str, res: ModelResults, inp: ModelInputs):
    """Generate branded Excel workbook with all model outputs."""
    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════════════════
    # Sheet 1: Summary
    # ══════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = "00B4A0"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 28

    # Header
    ws.merge_cells("A1:B1")
    c = ws.cell(row=1, column=1, value="AMPYR DISTRIBUTED ENERGY")
    c.font = Font(color="FFFFFF", bold=True, size=14)
    c.fill = _NAVY_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=1, column=2).fill = _NAVY_FILL
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:B2")
    c2 = ws.cell(row=2, column=1, value="Solar PPA Pricing Report")
    c2.font = Font(color="00B4A0", size=11)
    c2.fill = PatternFill(start_color="122240", end_color="122240", fill_type="solid")
    ws.cell(row=2, column=2).fill = PatternFill(start_color="122240", end_color="122240", fill_type="solid")

    ws.merge_cells("A3:B3")
    c3 = ws.cell(row=3, column=1, value=inp.project_name)
    c3.font = Font(bold=True, size=12)

    row = 5

    # Warnings
    if res.warnings:
        row = _write_kv_section(ws, row, "⚠ Warnings",
                                [(w, "") for w in res.warnings])

    # Key Results
    row = _write_kv_section(ws, row, "Key Results", [
        ("Solve Method", res.solve_method),
        ("Year 1 PPA Rate", f"{res.ppa_rate_year1_pence:.2f} p/kWh"),
        ("Development Fee", f"£{res.dev_fee_total_solved:,.0f} (£{res.dev_fee_per_kwp_solved:,.1f}/kWp)"),
        ("Levered IRR", f"{res.levered_irr * 100:.2f}%"),
        ("Target IRR", f"{res.target_irr * 100:.2f}%"),
        ("Unlevered IRR", f"{res.unlevered_irr * 100:.2f}%"),
        ("Cash Multiple", f"{res.cash_multiple:.2f}x"),
        ("Payback Period", f"Year {res.payback_year}"),
        ("Average DSCR", f"{res.average_dscr:.2f}x"),
    ])

    # Key Dates
    row = _write_kv_section(ws, row, "Key Dates", [
        ("Financial Close", res.financial_close.strftime("%d %b %Y")),
        ("Construction Start", res.construction_start.strftime("%d %b %Y")),
        ("COD", res.cod.strftime("%d %b %Y")),
        ("Operations End", res.operations_end.strftime("%d %b %Y")),
        ("First Indexation", res.first_indexation_date.strftime("%d %b %Y")),
        ("Senior Debt Maturity", res.debt_end.strftime("%d %b %Y")),
    ])

    # Capex
    row = _write_kv_section(ws, row, "Capital Expenditure", [
        ("EPC Total", f"£{res.epc_total:,.0f}"),
        ("Insurance (Construction)", f"£{res.insurance_construction:,.0f}"),
        ("Transaction Costs", f"£{res.transaction_costs:,.0f}"),
        ("Development Fee", f"£{res.development_fee_total:,.0f}"),
        ("Management Company Fee", f"£{res.management_company_fee:,.0f}"),
        ("Stamp Duty", f"£{res.stamp_duty:,.0f}"),
        ("Contingency", f"£{res.contingency:,.0f}"),
        ("Total Capex", f"£{res.total_capex:,.0f}"),
    ])

    # Funding
    row = _write_kv_section(ws, row, "Funding Structure", [
        ("Senior Debt", f"£{res.senior_debt_amount:,.0f} ({inp.debt_gearing * 100:.0f}%)"),
        ("SHL", f"£{res.shl_amount:,.0f}"),
        ("Pure Equity", f"£{res.pure_equity:,.0f}"),
    ])

    # Customer Savings
    row = _write_kv_section(ws, row, "Customer Savings", [
        ("Year 1 Saving", f"£{res.year1_saving:,.0f}"),
        ("Average Annual Saving", f"£{res.avg_annual_saving:,.0f}"),
        ("Total Lifetime Saving", f"£{res.total_lifetime_saving:,.0f}"),
        ("Avg Lifetime PPA Rate", f"{res.avg_lifetime_ppa_rate:.2f} p/kWh"),
        ("Avg Lifetime Grid Rate", f"{res.avg_lifetime_grid_rate:.2f} p/kWh"),
    ])

    # ══════════════════════════════════════════════════════════════════
    # Sheet 2: Annual Cashflows
    # ══════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Cashflows")
    ws2.sheet_properties.tabColor = "0B1A2E"

    headers = ["Year", "Generation (MWh)", "PPA Rate (p/kWh)", "Revenue (£)",
               "Total Opex (£)", "EBITDA (£)", "Capital Allow. (£)",
               "Tax (£)", "CFADS (£)", "Debt Interest (£)", "Debt Principal (£)",
               "Debt Service (£)", "DSCR", "FCF to Equity (£)",
               "SHL Interest (£)", "SHL Repay (£)", "Equity Dist (£)",
               "Cum. Distributions (£)", "Debt Balance (£)", "SHL Balance (£)"]

    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _NAVY_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.column_dimensions[get_column_letter(col)].width = 16

    for i, cf in enumerate(res.cashflows):
        r = i + 2
        vals = [
            cf.year, cf.generation_kwh / 1000, cf.ppa_price_pence,
            cf.revenue, cf.total_opex, cf.ebitda, cf.capital_allowance,
            cf.tax, cf.cfads, cf.debt_interest, cf.debt_principal,
            cf.debt_service, cf.dscr if cf.dscr > 0 else None,
            cf.fcf_to_equity, cf.shl_interest, cf.shl_repayment,
            cf.equity_distribution, cf.cumulative_distributions,
            cf.debt_balance_end, cf.shl_balance_end,
        ]
        for col, v in enumerate(vals, 1):
            cell = ws2.cell(row=r, column=col, value=v)
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER
            if r % 2 == 0:
                cell.fill = _LIGHT_FILL
            # Number formatting
            if col >= 4 and col != 13:
                cell.number_format = '#,##0'
            elif col == 3:
                cell.number_format = '0.00'
            elif col == 13 and v is not None:
                cell.number_format = '0.00'

    # ══════════════════════════════════════════════════════════════════
    # Sheet 3: Customer View
    # ══════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Customer View")
    ws3.sheet_properties.tabColor = "00B4A0"

    cust_headers = ["Year", "Generation (MWh)", "PPA Rate (p/kWh)", "PPA Cost (£)",
                    "Grid Rate (p/kWh)", "Grid Cost (£)", "Annual Saving (£)",
                    "Cumulative Saving (£)", "Buyout Value (£)"]

    for col, h in enumerate(cust_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _TEAL_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws3.column_dimensions[get_column_letter(col)].width = 18

    cum_saving = 0
    for i, cf in enumerate(res.cashflows):
        r = i + 2
        cum_saving += cf.customer_saving
        vals = [
            cf.year, cf.generation_kwh / 1000, cf.ppa_price_pence,
            cf.revenue, cf.grid_price_pence, cf.grid_cost,
            cf.customer_saving, cum_saving, cf.buyout_value,
        ]
        for col, v in enumerate(vals, 1):
            cell = ws3.cell(row=r, column=col, value=v)
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER
            if r % 2 == 0:
                cell.fill = _LIGHT_FILL
            if col in (3, 5):
                cell.number_format = '0.00'
            elif col >= 4:
                cell.number_format = '#,##0'

    # ══════════════════════════════════════════════════════════════════
    # Sheet 4: Inputs
    # ══════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Inputs")
    ws4.sheet_properties.tabColor = "8899AA"
    ws4.column_dimensions["A"].width = 35
    ws4.column_dimensions["B"].width = 25

    ws4.cell(row=1, column=1, value="Input Parameter").font = _HEADER_FONT
    ws4.cell(row=1, column=1).fill = _NAVY_FILL
    ws4.cell(row=1, column=2, value="Value").font = _HEADER_FONT
    ws4.cell(row=1, column=2).fill = _NAVY_FILL

    input_pairs = [
        ("Project Name", inp.project_name),
        ("COD Date", inp.cod_date.strftime("%d/%m/%Y")),
        ("Construction Period", f"{inp.effective_construction_months()} months"),
        ("Credit Score", inp.credit_score),
        ("Project Size", f"{inp.project_size_kwp:,.1f} kWp"),
        ("Specific Yield", f"{inp.specific_yield:,.0f} kWh/kWp"),
        ("EPC Price", f"£{inp.epc_per_kwp:,.0f}/kWp"),
        ("PPA Length", f"{inp.ppa_length_years} years"),
        ("PPA Indexation", inp.ppa_indexation),
        ("Assumed Inflation", f"{inp.assumed_inflation_rate * 100:.1f}%"),
        ("Fixed Inflation", f"{inp.fixed_inflation_rate * 100:.1f}%"),
        ("Take or Pay", f"{inp.take_or_pay_pct * 100:.0f}%"),
        ("Take or Pay Method", inp.take_or_pay_method),
        ("Solve Method", inp.solve_method),
        ("PPA Price Input", f"{inp.ppa_price_input_pence:.1f} p/kWh"),
        ("Dev Fee Input", f"£{inp.dev_fee_per_kwp:,.1f}/kWp"),
        ("Grid Price", f"{inp.grid_price_pence:.1f} p/kWh"),
        ("Grid Escalation", f"{inp.grid_price_escalation * 100:.1f}%"),
        ("Debt Gearing", f"{inp.debt_gearing * 100:.0f}%"),
        ("Debt Interest Rate", f"{inp.debt_interest_rate * 100:.1f}%"),
        ("Debt Tenor", f"{inp.debt_tenor_years} years"),
        ("Repayment Method", inp.repayment_method),
        ("DSCR Target", f"{inp.dscr_target:.2f}x"),
        ("Equity LTV", f"{inp.equity_ltv * 100:.0f}%"),
        ("SHL Coupon", f"{inp.shl_coupon * 100:.1f}%"),
        ("SHL Tenor", f"{inp.shl_tenor_years} years"),
        ("Stamp Duty", f"{inp.stamp_duty_pct * 100:.1f}%"),
        ("Contingency", f"{inp.contingency_pct * 100:.1f}%"),
        ("Management Fee", f"{inp.management_company_fee_pct * 100:.1f}%"),
        ("Tax Rate", f"{inp.tax_rate * 100:.0f}%"),
        ("CA Method", inp.capital_allowance_method),
        ("CA Rate", f"{inp.capital_allowance_rate * 100:.1f}%"),
        ("Degradation", f"{inp.annual_degradation * 100:.2f}%"),
    ]

    for i, (k, v) in enumerate(input_pairs):
        r = i + 2
        ws4.cell(row=r, column=1, value=k).font = Font(bold=True, size=9)
        ws4.cell(row=r, column=2, value=str(v)).font = _BODY_FONT
        if r % 2 == 0:
            ws4.cell(row=r, column=1).fill = _LIGHT_FILL
            ws4.cell(row=r, column=2).fill = _LIGHT_FILL

    wb.save(filepath)
